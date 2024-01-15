# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import sys
from io import BytesIO


class HuExcelParser:
    def __call__(self, fnm):
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(fnm))
        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            ti = list(rows[0])
            for r in list(rows[1:]):
                l = []
                for i,c in enumerate(r):
                    if not c.value:continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("："  if t else "") + str(c.value)
                    l.append(t)
                l = "; ".join(l)
                if sheetname.lower().find("sheet") <0: l += " ——"+sheetname
                res.append(l)
        return res


if __name__ == "__main__":
    psr = HuExcelParser()
    psr(sys.argv[1])
